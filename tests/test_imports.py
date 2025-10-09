from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from src.app.db import connect, init_db, record_print
from src.app.imports import (
    ValidationParseResult,
    apply_validation_updates,
    build_ddn_lookup_from_rows,
    import_already_printed_csv,
    load_last_import,
    parse_validation_workbook,
    persist_last_import,
)
from src.app.io_utils import lire_tableau


def test_build_ddn_lookup_handles_conflicts():
    rows = [
        {"Nom": "Dup", "Prénom": "Test", "Date_de_naissance": "01/01/2000"},
        {"Nom": "Dup", "Prénom": "Test", "Date_de_naissance": "02/02/2000"},
        {"Nom": "Unique", "Prénom": "One", "Date_de_naissance": "03/03/2000"},
        {"Nom": "dùpônt", "Prénom": "álix", "Date_de_naissance": "04/04/2000"},
        {"Nom": "", "Prénom": "", "Date_de_naissance": ""},
    ]

    lookup = build_ddn_lookup_from_rows(rows)
    assert lookup[("unique", "one")] == "03/03/2000"
    assert lookup[("dup", "test")] is None
    assert lookup[("dupont", "alix")] == "04/04/2000"


def test_import_already_printed_csv_resolves_ddn(tmp_path):
    db_path = tmp_path / "app.db"
    init_db(db_path)

    with connect(db_path) as cn:
        # Unique DDN for Beta
        record_print(cn, "Beta", "User", "10/10/2010", "01/01/2024", zpl=None, status="printed")
        # Multiple DDN for Gamma to trigger the "ambiguous" path
        record_print(cn, "Gamma", "User", "11/11/2011", "01/01/2024", zpl=None, status="printed")
        record_print(cn, "Gamma", "User", "22/02/2012", "02/02/2024", zpl=None, status="printed")

    csv_path = tmp_path / "import.csv"
    csv_path.write_text("Alpha;Test\nBeta;User\nGamma;User\n;\n", encoding="utf-8")

    rows_lookup = build_ddn_lookup_from_rows(
        [{"Nom": "Alpha", "Prénom": "Test", "Date_de_naissance": "05/05/2005"}]
    )

    imported, skipped = import_already_printed_csv(
        csv_path,
        "31/12/2025",
        rows_ddn_lookup=rows_lookup,
        db_path=db_path,
    )

    assert imported == 3
    assert skipped == 1

    with connect(db_path) as cn:
        rows = cn.execute(
            "SELECT nom, prenom, ddn FROM prints WHERE expire=? ORDER BY nom",
            ("31/12/2025",),
        ).fetchall()

    assert [tuple(row) for row in rows] == [
        ("ALPHA", "TEST", "05/05/2005"),
        ("BETA", "USER", "10/10/2010"),
        ("GAMMA", "USER", ""),
    ]


def test_import_normalizes_names(tmp_path):
    db_path = tmp_path / "app.db"
    init_db(db_path)

    csv_path = tmp_path / "import.csv"
    csv_path.write_text("dùpônt;álix\n", encoding="utf-8")

    rows_lookup = build_ddn_lookup_from_rows(
        [{"Nom": "dùpônt", "Prénom": "álix", "Date_de_naissance": "01/01/2000"}]
    )

    imported, skipped = import_already_printed_csv(
        csv_path,
        "31/12/2025",
        rows_ddn_lookup=rows_lookup,
        db_path=db_path,
    )

    assert imported == 1
    assert skipped == 0

    with connect(db_path) as cn:
        row = cn.execute(
            "SELECT nom, prenom, ddn FROM prints WHERE expire=?",
            ("31/12/2025",),
        ).fetchone()

    assert tuple(row) == ("DUPONT", "ALIX", "01/01/2000")


def test_parse_validation_workbook_reformats(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Durand Léa", "Confirmé le 01/06/2024 par Marie Curie", "", "", "", ""])
    ws.append(["13 ans", "Donnée inutile", "", "", "", ""])
    ws.append(["Membre", "", "", "", "", "Montant : 19,50 €"])
    ws.append(["Inutile", "", "", "", "", "Encore"])
    ws.append(["", "", "", "", "", ""])
    ws.append(["Martin Paul", "", "", "", "", ""])
    ws.append(["", "Confirmé le 02-06-2024 par Toto", "", "", "", ""])
    ws.append(["", "", "", "", "", "Montant 20 €"])
    ws.append(["", "Ligne en trop", "", "", "", ""])

    ws["A1"].hyperlink = "https://example.com"
    ws["A1"].style = "Hyperlink"
    ws["A6"].hyperlink = "https://example.com/member"
    ws["A6"].style = "Hyperlink"

    path = tmp_path / "validation.xlsx"
    wb.save(path)

    result = parse_validation_workbook(path, export_dir=tmp_path, now=datetime(2024, 6, 7, 15, 30))

    assert isinstance(result, ValidationParseResult)
    assert len(result.rows) == 2

    first, second = result.rows
    assert first["Nom"] == "DURAND"
    assert first["Prénom"] == "LEA"
    assert first["Montant"] == "19.50"
    assert first["Validation_confirmee_le"] == "01/06/2024"
    assert first["Validation_confirmee_par"] == "MARIE CURIE"

    assert second["Nom"] == "MARTIN"
    assert second["Prénom"] == "PAUL"
    assert second["Montant"] == "20.00"
    assert second["Validation_confirmee_le"] == "02/06/2024"
    assert second["Validation_confirmee_par"] == "TOTO"

    assert result.export_path is not None
    assert result.export_path.name == "FichierValidation_20240607-153000.xlsx"
    exported = pd.read_excel(result.export_path)
    assert exported.shape[0] == 2
    assert list(exported.columns) == [
        "Nom",
        "Prénom",
        "Date_de_naissance",
        "Expire_le",
        "Email",
        "Montant",
        "ErreurValide",
        "Validation_confirmee_le",
        "Validation_confirmee_par",
    ]


def test_parse_validation_workbook_handles_multiline_blocks(tmp_path):
    data = [
        ["CARDON Elise", "Confirmé le 01/06/2024 par AUBER Yann", "", "", "", ""],
        ["13 ans", "", "", "", "", ""],
        ["Membre", "", "", "", "", "Montant : 19,50 €"],
        ["Donnée inutile", "", "", "", "", "Ignorer"],
        ["", "", "", "", "", ""],
        ["DUPONT Marc", "", "", "", "", ""],
        ["", "Confirmé le 02-06-2024 par BOB Martin", "", "", "", ""],
        ["Membre", "", "", "", "", "Montant 20 €"],
    ]
    path = tmp_path / "validation_raw.xlsx"
    pd.DataFrame(data).to_excel(path, index=False, header=False)

    result = parse_validation_workbook(path, export_dir=tmp_path, now=datetime(2024, 6, 7, 8, 15))

    assert isinstance(result, ValidationParseResult)
    assert len(result.rows) == 2

    first, second = result.rows
    assert first["Nom"] == "CARDON"
    assert first["Prénom"] == "ELISE"
    assert first["Montant"] == "19.50"
    assert first["Validation_confirmee_le"] == "01/06/2024"
    assert first["Validation_confirmee_par"] == "AUBER YANN"

    assert second["Nom"] == "DUPONT"
    assert second["Prénom"] == "MARC"
    assert second["Montant"] == "20.00"
    assert second["Validation_confirmee_par"] == "BOB MARTIN"

    assert result.export_path is not None
    assert result.export_path.name == "FichierValidation_20240607-081500.xlsx"
    exported = pd.read_excel(result.export_path)
    assert exported.shape[0] == 2
    assert list(exported.columns) == [
        "Nom",
        "Prénom",
        "Date_de_naissance",
        "Expire_le",
        "Email",
        "Montant",
        "ErreurValide",
        "Validation_confirmee_le",
        "Validation_confirmee_par",
    ]


def test_apply_validation_updates_updates_and_adds():
    existing = [
        {
            "Nom": "ALPHA",
            "Prénom": "TEST",
            "Date_de_naissance": "01/01/2000",
            "Expire_le": "31/12/2025",
            "Email": "",
            "Montant": "",
            "ErreurValide": "",
            "Validation_confirmee_le": "",
            "Validation_confirmee_par": "",
            "Derniere": "2024-01-01T12:00:00",
            "Compteur": 2,
        }
    ]
    updates = [
        {
            "Nom": "Alpha",
            "Prénom": "Test",
            "Date_de_naissance": "01/01/2000",
            "Email": "alpha@example.com",
            "ErreurValide": "Yes",
            "Validation_confirmee_le": "02/02/2024",
            "Validation_confirmee_par": "VALIDATOR",
        },
        {
            "Nom": "Beta",
            "Prénom": "User",
            "Date_de_naissance": "",
            "Expire_le": "31/12/2026",
            "Montant": "42",
            "ErreurValide": "No",
            "Validation_confirmee_le": "03/03/2024",
            "Validation_confirmee_par": "CHECKER",
        },
    ]

    merged, updated, added = apply_validation_updates(existing, updates)

    assert updated == 1
    assert added == 1
    assert merged[0]["Email"] == "alpha@example.com"
    assert merged[0]["ErreurValide"] == "Yes"
    assert merged[0]["Validation_confirmee_le"] == "02/02/2024"
    assert merged[0]["Validation_confirmee_par"] == "VALIDATOR"
    assert any(r["Nom"] == "BETA" and r["Prénom"] == "USER" for r in merged)
    new_row = next(r for r in merged if r["Nom"] == "BETA")
    assert new_row["Compteur"] == 0
    assert new_row["Derniere"] == ""
    assert new_row["Validation_confirmee_par"] == "CHECKER"


def _write_sample_csv(path: Path) -> None:
    path.write_text(
        "Ignoré\nIgnoré\nIgnoré\n"
        "Nom,Prénom,Date_de_naissance,Expire_le,Email,Montant,ErreurValide\n"
        "Alpha,Test,01/01/2000,31/12/2025,alpha@example.com,123.45,Yes\n",
        encoding="utf-8",
    )


def _write_csv_without_optional_columns(path: Path) -> None:
    path.write_text(
        "Ignoré\nIgnoré\nIgnoré\n"
        "Nom,Prénom,Date_de_naissance,Expire_le,Email\n"
        "Beta,User,02/02/2002,31/12/2025,beta@example.com\n",
        encoding="utf-8",
    )


def test_lire_tableau_adds_missing_optional_columns(tmp_path):
    sample = tmp_path / "import.csv"
    _write_csv_without_optional_columns(sample)

    df = lire_tableau(sample)

    assert list(df.columns) == [
        "Nom",
        "Prénom",
        "Date_de_naissance",
        "Expire_le",
        "Email",
        "Montant",
        "ErreurValide",
    ]
    assert df.iloc[0]["Nom"] == "BETA"
    assert df.iloc[0]["Montant"] == ""
    assert df.iloc[0]["ErreurValide"] == ""


def test_persist_and_load_last_import(tmp_path, monkeypatch):
    from src.app import imports as imports_mod

    source = tmp_path / "import.csv"
    _write_sample_csv(source)

    cache_dir = tmp_path / "cache"
    metadata_file = tmp_path / "meta.json"

    monkeypatch.setattr(imports_mod, "LAST_IMPORT_DIR", cache_dir)
    monkeypatch.setattr(imports_mod, "LAST_IMPORT_METADATA", metadata_file)

    metadata = persist_last_import(source)

    assert metadata_file.exists()
    assert Path(metadata["cached_path"]).exists()
    assert metadata["source_name"] == source.name
    assert metadata["cached_name"] == source.name
    assert "stored_at" in metadata

    rows, loaded_metadata = load_last_import()

    assert len(rows) == 1
    assert rows[0]["Nom"] == "ALPHA"
    assert rows[0]["Email"] == "alpha@example.com"
    assert rows[0]["Montant"] == "123.45"
    assert rows[0]["ErreurValide"] == "Yes"
    assert loaded_metadata["cached_name"] == source.name
